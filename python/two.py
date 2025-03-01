from openpyxl import Workbook

# Создаем новую рабочую книгу
wb = Workbook()
ws = wb.active

# Заголовки для таблицы умножения
ws.cell(row=1, column=1, value="")  # Пустая ячейка в углу
for i in range(2, 10):
    ws.cell(row=1, column=i, value=i)  # Заполняем верхнюю строку числами от 2 до 9

for i in range(2, 10):  # Строки для чисел от 2 до 9
    ws.cell(row=i, column=1, value=i)  # Левый столбец с числами от 2 до 9
    for j in range(2, 10):  # Столбцы для чисел от 2 до 9
        ws.cell(row=i, column=j, value=ws.cell(row=i, column=1).value * ws.cell(row=1, column=j).value)

# Сохраняем таблицу в файл
wb.save("таблица_умножения.xlsx")

print("Таблица умножения успешно создана!")